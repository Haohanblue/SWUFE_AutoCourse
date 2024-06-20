import tkinter as tk
import ttkbootstrap as ttk
from tkinter import messagebox, scrolledtext, simpledialog
from ttkbootstrap import Style
from ttkbootstrap.constants import *
import requests
import re
import time
from bs4 import BeautifulSoup
import json
from login import login
from xkqk import xkqk
import os
import subprocess
from clear import clear
import ast
import pandas as pd
from openpyxl import load_workbook
monitor_process = None
is_auto_submitting = False
jianting_process = None


def open_config_window_if_empty():
    try:
        with open("config.json", "r",encoding='utf-8') as json_file:
            config_data = json.load(json_file)
        if not config_data.get("XH") or not config_data.get("PWD") or not config_data.get("TO_EMAIL") or not config_data.get("URL"):
            open_config_window()
            messagebox.showinfo("提示","请先设置配置信息")
    except FileNotFoundError:
        open_config_window()


def on_close():
    global monitor_process
    if monitor_process and monitor_process.poll() is None:
        monitor_process.terminate()
        monitor_process.wait()
    root.destroy()

def get_login_url():
    max_try = 0
    while max_try < 3 :
        try:
            config = None
            with open('config.json', 'r',encoding='utf-8') as config_file:
                config = json.load(config_file)
            if config:
                login_url = login(config["URL"], config["XH"], config["PWD"])
                if "string indices must be integers" in login_url:
                    show_log("Cookies已经过期，正在重新获取")
                else:
                    show_log(f"登录后的页面URL: {login_url}", True)
                    break
        except Exception as e:
            show_alert(f"获取登录URL失败：{str(e)}")
            clear()
            time.sleep(1)
        max_try += 1



def get_current_selection():
    try:
        config = None
        with open('./config.json', 'r',encoding='utf-8') as config_file:
            config = json.load(config_file)
        if config:
            result = xkqk(config["MAINURL"])
            i = 0
            show_log("当前选课情况:", True)
            for item in result:
                i += 1
                show_log(f'{i}.{item}', False)

    except Exception as e:
        show_alert(f"获取选课情况失败：{str(e)}")

def start_stop_monitor():
    global jianting_process
    if jianting_process is None:
        try:
            jianting_process = subprocess.Popen(["jianting.exe"], text=True, creationflags=subprocess.CREATE_NEW_PROCESS_GROUP)
            show_log("监听执行中", True)
        except Exception as e:
            show_alert(f"启动监听失败：{str(e)}")
    else:
        try:
            jianting_process.terminate()
            jianting_process.wait()
            show_alert("已经关闭监听")
            jianting_process = None
        except Exception as e:
            show_alert(f"结束监听失败：{str(e)}")

def open_config_window():
    config_window = tk.Toplevel(main_frame)
    config_window.title("修改配置文件")
    config_window.attributes('-topmost',True)
    labels = ["学号(必填):", "密码(必填):", "邮箱:", "教务系统URL(必填):","有容量的xkkh(必填):","系统主页","Cookies","VIEWSTATE"]
    names = ['XH', 'PWD', 'TO_EMAIL', 'URL',"XKKH","MAINURL","COOKIES","VIEWSTATE"]
    entries = [ttk.Entry(config_window, width=40) for _ in range(len(labels))]

    try:
        with open("config.json", "r",encoding='utf-8') as json_file:
            config_data = json.load(json_file)
            for entry, name in zip(entries, names):
                entry.insert(0, config_data.get(name, ""))
    except FileNotFoundError:
        pass

    for label, entry in zip(labels, entries):
        ttk.Label(config_window, text=label).grid(sticky="w", padx=10, pady=5)
        entry.grid(sticky="w", padx=10, pady=5)

    def save_config():
        try:
            with open("config.json", "r",encoding='utf-8') as json_file:
                config_data = json.load(json_file)

            for name, entry in zip(names, entries):
                config_data[name] = entry.get()

            with open("config.json", "w") as json_file:
                json.dump(config_data, json_file, indent=4)

            config_window.destroy()
            messagebox.showinfo("成功", "配置文件已更新")
        except Exception as e:
            messagebox.showerror("错误", f"无法保存配置文件：{str(e)}")

    button_save = ttk.Button(config_window, text="保存", command=save_config,bootstyle="outline-success")
    button_save.grid(row=2*len(labels), column=0, columnspan=4, pady=10)

def show_log(message, is_time=False):
    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    log_message = f"[{current_time}] {message}\n" if is_time else f"{message}\n"
    text_result.config(state=tk.NORMAL)
    text_result.insert(tk.END, log_message)
    text_result.yview(tk.END)
    text_result.config(state=tk.DISABLED)

def show_alert(alert_message):
    show_log(alert_message, True)

def extract_alert(response_text):
    soup = BeautifulSoup(response_text, 'html.parser')
    try:
        alert_message = soup.find('script', language="javascript").text
        alert_content = re.search(r"alert\('(.*?)'\);", alert_message)
        if alert_content:
            return alert_content.group(1)
        else:
            message = soup.find_all('script')[-1].text
            alert_content = re.search(r"alert\('(.*?)'\);", message)
            if alert_content:
                content = alert_content.group(1)
                return content
            else:
                return "出现错误"
    except Exception as e:
        alert_message = soup.find('title').text
        return alert_message

def submit_form():
    server_script_url = entry_server_script_url.get()
    print(server_script_url)
    is_textbook_ordered = radio_var.get()
    course_codes = [entry.get() for entry in entry_course_codes]
    course_VIEWSTATE = [entry.get() for entry in entry_course_VIEWSTATE]

    config = None
    with open('config.json', 'r',encoding='utf-8') as config_file:
        config = json.load(config_file)
    if config:
        cookies = config["COOKIES"]
        VIEWSTATE = config["VIEWSTATE"]
        if type(cookies) == str:
            cookies = ast.literal_eval(cookies)
    for i, course_code in enumerate(course_codes, start=1):
        if course_code != "":
            print(course_code,type(course_code))
            if course_VIEWSTATE[i-1] == "":
                course_VIEWSTATE[i-1] = VIEWSTATE
            form_data = {
                '__EVENTTARGET': 'Button1',
                # '__EVENTARGUMENT': '',
                 '__VIEWSTATEGENERATOR': 'AC27D4D4',
                '__VIEWSTATE': course_VIEWSTATE[i-1],
                'xkkh': course_code,
                'RadioButtonList1': is_textbook_ordered,
                  }
            headers = {
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Encoding': 'gzip, deflate',
                'Accept-Language': 'zh-CN,zh;q=0.8',
                'Cache-Control': 'max-age=0',
                'Connection': 'keep-alive',
                'Content-Type': 'application/x-www-form-urlencoded',
                'Cookie': f'route={cookies["route"]}',
                'Host': 'xk.swufe.edu.cn',
                'Origin': 'http://xk.swufe.edu.cn',
                'Referer': server_script_url,
                'Upgrade-Insecure-Requests': '1',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
            }
            session = requests.Session()
            # print(cookies)
            #
            # session.cookies.set('route', f'{cookies["route"]}', path='/')
            #
            try:
                session.cookies.update(cookies)
                session.headers.update(headers)
                response = session.post(server_script_url, data=form_data)
                if response.status_code == 200:
                    # print(response.text)
                    content = extract_alert(response.text)
                    # print(server_script_url)
                    # print(session.headers)
                    # print("responseHeader:",response.headers)
                    # print("responseURL:",response.url)
                    # print("responseCookies:",response.cookies)
                    alert_message = f"第{i}个：" + content
                    show_log(alert_message, True)
                    if "教师" in content or "重修" in content:
                        with open('config.json', 'r') as config_file:
                            config = json.load(config_file)
                        if config:
                            show_log("已经自动重新获取选课URL", True)
                            login(config["URL"],config["XH"],config["PWD"])
                            auto_fill()
                else:
                    show_alert(f"请求失败，状态码: {response.status_code}")
            except Exception as e:
                print("出错啦")
                show_log(e)
                if "string indices must be integers" in str(e):
                    show_log("Cookies已经过期，正在重新获取")
                    get_login_url()

def auto_submit_form():
    global is_auto_submitting
    if is_auto_submitting:
        submit_form()
        main_frame.after(int(float(custom_interval.get()) * 1000), auto_submit_form)

def toggle_auto_submitting():
    global is_auto_submitting
    is_auto_submitting = not is_auto_submitting
    if is_auto_submitting:
        show_log("已开始自动抢课", True)
        auto_submit_form()
    else:
        show_log("已停止自动抢课", True)

def auto_fill():
    try:
        with open('config.json', 'r',encoding='utf-8') as config_file:
            config = json.load(config_file)
        if config:
            baseurl=config["MAINURL"]
            xkkh = config["XKKH"]
            xh = config["XH"]
            xkurl = re.match(r"(http?://[^/]+/[^/]+/)", baseurl).group(1)
            result = xkurl+"xsxjs.aspx?xkkh="+xkkh+"&xh="+xh
            VIEWSTATE = config["VIEWSTATE"]
            if VIEWSTATE =="":
                cookies = config["COOKIES"]
                if type(cookies) == str:
                    cookies = ast.literal_eval(cookies)
                session = requests.Session()
                headers = {
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                    'Accept-Encoding': 'gzip,deflate',
                    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
                    'Connectiong': 'keep-alive',
                    'Cookie':f'route={cookies["route"]}',
                    'Host': 'xk.swufe.edu.cn',
                    'Upgrade-Insecure-Requests': '1',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
                }
                session.cookies.update(cookies)
                session.headers.update(headers)
                print(result)
                response = session.get(result,headers=headers,timeout=3)
                if response.status_code == 200:
                    show_log("检测到VIEWSTATE为空，正在自动获取VIEWSTATE", True)
                    content = response.text
                    soup = BeautifulSoup(content, 'html.parser')
                    VIEWSTATE = soup.find('input', {'name': '__VIEWSTATE'}).get('value')
                    #将VIEWSTATE写入config.json
                    with open("config.json", "r") as json_file:
                        config_data = json.load(json_file)
                        config_data['VIEWSTATE'] = VIEWSTATE
                    with open("config.json", "w") as json_file:
                        json.dump(config_data, json_file, indent=4)
                        show_log("填充VIEWSTATE完毕，如需自定义请修改配置文件", True)
            if result == "三秒防刷":
                show_log(result, True)
            else:
                entry_server_script_url.delete(0, tk.END)
                entry_server_script_url.insert(0, result)
                show_log(f"获取到的URL为:\n{result}", True)
    except Exception as e:
        show_log(e, True)
        entry_server_script_url.delete(0, tk.END)
        show_log(0, "自动填充失败")

def add_course_entry():
    new_entry = ttk.Entry(main_frame, width=40)
    new_entry.grid(row=len(entry_course_codes) + 3, column=1, padx=10, pady=5, sticky="w")
    new_entry2 = ttk.Entry(main_frame, width=30)
    new_entry2.grid(row=len(entry_course_codes) + 3, column=3, padx=10, pady=5, sticky="w")

    entry_course_codes.append(new_entry)
    entry_course_VIEWSTATE.append(new_entry2)
    if len(entry_course_codes)>2:
        adjust_layout()

def remove_course_entry():
    if len(entry_course_codes) > 1:
        entry_to_remove = entry_course_codes.pop()
        entry_to_remove.destroy()
        entry_to_remove2 = entry_course_VIEWSTATE.pop()
        entry_to_remove2.destroy()

        if len(entry_course_codes) > 1:
            adjust_layout()

def get_logged_in_url():
    try:
        server_script_url = entry_server_script_url.get()
        logged_in_url = login(server_script_url)
        show_log(f"登录后页面URL: {logged_in_url}", True)
    except Exception as e:
        show_alert(f"获取登录后页面URL失败：{str(e)}")


def open_selection_editor():
    global active_listbox
    active_listbox = None
    selection_editor_window = tk.Toplevel(main_frame)
    selection_editor_window.title("选课代码编辑器")

    # 读取Excel文件内容
    file_path = "选课代码.xlsx"
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    label_column1 = ttk.Label(selection_editor_window, text="选课代码")
    label_column1.grid(row=0, column=0, padx=10, pady=10, sticky="w")

    listbox_column1 = tk.Listbox(selection_editor_window, selectmode=tk.SINGLE, height=20, width=25)
    listbox_column1.grid(row=1, column=0, padx=10, pady=10)

    label_column2 = ttk.Label(selection_editor_window, text="VIEWSTATE")
    label_column2.grid(row=0, column=1, padx=10, pady=10, sticky="w")
    listbox_column2 = tk.Listbox(selection_editor_window, selectmode=tk.SINGLE, height=20, width=25)
    listbox_column2.grid(row=1, column=1, padx=10, pady=10)

    label_column3 = ttk.Label(selection_editor_window, text="课程名称")
    label_column3.grid(row=0, column=2, padx=10, pady=10, sticky="w")
    listbox_column3 = tk.Listbox(selection_editor_window, selectmode=tk.SINGLE, height=20, width=25)
    listbox_column3.grid(row=1, column=2, padx=10, pady=10)

    # 填充Listbox
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=3, values_only=True):
        listbox_column1.insert(tk.END, row[0])
        listbox_column2.insert(tk.END, row[1])
        listbox_column3.insert(tk.END, row[2])

    def on_select(event):
        # 事件处理，当前未使用，但可以在此处理选择事件
        pass

    def edit_column1(event):
        selected_index = listbox_column1.curselection()
        if selected_index:
            current_value = listbox_column1.get(selected_index[0])
            new_value = simpledialog.askstring("编辑数据", "修改数据为:", initialvalue=current_value)
            if new_value:
                listbox_column1.delete(selected_index[0])
                listbox_column1.insert(selected_index[0], new_value)
                sheet.cell(row=selected_index[0] + 2, column=1, value=new_value)

    def edit_column2(event):
        selected_index = listbox_column2.curselection()
        if selected_index:
            current_value = listbox_column2.get(selected_index[0])
            new_value = simpledialog.askstring("编辑数据", "修改数据为:", initialvalue=current_value)
            if new_value:
                listbox_column2.delete(selected_index[0])
                listbox_column2.insert(selected_index[0], new_value)
                sheet.cell(row=selected_index[0] + 2, column=2, value=new_value)
    def edit_column3(event):
        selected_index = listbox_column3.curselection()
        if selected_index:
            current_value = listbox_column3.get(selected_index[0])
            new_value = simpledialog.askstring("编辑数据", "修改数据为:", initialvalue=current_value)
            if new_value:
                listbox_column3.delete(selected_index[0])
                listbox_column3.insert(selected_index[0], new_value)
                sheet.cell(row=selected_index[0] + 2, column=3, value=new_value)

    def save_and_close():
        workbook.save(filename=file_path)
        selection_editor_window.destroy()


    def add_new_code():
        new_code = simpledialog.askstring("新增选课代码", "输入新的选课代码:")
        if new_code:
            listbox_column1.insert(tk.END, new_code)
            # 添加数据到Excel文件
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=1, value=new_code)

    def add_new_viewstate():
        new_viewstate = simpledialog.askstring("新增VIEWSTATE", "输入新的VIEWSTATE:")
        if new_viewstate:
            listbox_column2.insert(tk.END, new_viewstate)
            # 更新Excel文件中相应的列
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=2, value=new_viewstate)
    def add_new_name():
        new_name = simpledialog.askstring("新增名称", "输入新的名称:")
        if new_name:
            listbox_column3.insert(tk.END, new_name)
            # 更新Excel文件中相应的列
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=3, value=new_name)
    def update_active_listbox(event):
        global active_listbox
        active_listbox = event.widget


    listbox_column1.bind("<Double-1>", edit_column1)  # 绑定双击事件到第一列
    listbox_column2.bind("<Double-1>", edit_column2)  # 绑定双击事件到第二列
    listbox_column3.bind("<Double-1>", edit_column3)  # 绑定双击事件到第三列

    # 为选课代码和VIEWSTATE分别添加按钮
    button_add_code = ttk.Button(selection_editor_window, text="添加新选课代码", command=add_new_code,bootstyle="outline-prime")
    button_add_viewstate = ttk.Button(selection_editor_window, text="添加新VIEWSTATE", command=add_new_viewstate,bootstyle="outline-prime")
    button_add_name = ttk.Button(selection_editor_window, text="添加新名称", command=add_new_name,bootstyle="outline-prime")

    button_add_code.grid(row=2, column=0, padx=2, pady=2)
    button_add_viewstate.grid(row=2, column=1, padx=2, pady=2)
    button_add_name.grid(row=2, column=2, padx=2, pady=2)


    button_save_close = ttk.Button(selection_editor_window, text="保存并关闭", command=save_and_close,bootstyle="outline-success")
    button_save_close.grid(row=4, column=0, columnspan=3, pady=1)


def clearlogin():
    clear()
    show_log("已清空MAINURL和COOKIES",True)
def load_selection_codes():
    try:
        file_path = "选课代码.xlsx"
        if os.path.exists(file_path):
            df = pd.read_excel('选课代码.xlsx')  # 假设第一列是我们需要的数据
            course_code = df["course_code"].dropna().astype(str).tolist()
            course_VIEWSTATE = df['VIEWSTATE'].fillna('').astype(str).tolist()
            for i, code in enumerate(course_code, start=1):
                if i <= len(entry_course_codes):
                    entry_course_codes[i - 1].delete(0, tk.END)
                    entry_course_codes[i - 1].insert(tk.END, code)
                    entry_course_VIEWSTATE[i - 1].delete(0, tk.END)
                    entry_course_VIEWSTATE[i - 1].insert(tk.END, course_VIEWSTATE[i - 1])
                else:
                    add_course_entry()
                    entry_course_codes[-1].delete(0, tk.END)
                    entry_course_codes[-1].insert(tk.END, code)
                    entry_course_VIEWSTATE[i - 1].delete(0, tk.END)
                    entry_course_VIEWSTATE[i - 1].insert(tk.END, course_VIEWSTATE[i - 1])
            show_alert("选课代码已全部填充完毕")
        else:
            show_alert("选课代码.xlsx 文件不存在")
    except Exception as e:
        show_alert(f"加载选课代码失败：{str(e)}")
def adjust_layout():
    for idx, entry in enumerate(entry_course_codes, start=3):
        entry.grid(row=idx, column=1, padx=10, pady=5, sticky="w")
        button_start_stop_monitor.grid(row=len(entry_course_codes) + 8, column=0, columnspan=1, pady=10)
        button_submit.grid(row=len(entry_course_codes) + 4, column=1, columnspan=2, pady=10)
        label_custom_interval.grid(row=len(entry_course_codes) + 3, column=0, padx=10, pady=5, sticky="w")
        custom_interval.grid(row=len(entry_course_codes) + 3, column=1, padx=10, pady=5, sticky="w")
        button_auto_submit.grid(row=len(entry_course_codes) + 4, column=0, columnspan=2, pady=5)
        button_get_selection.grid(row=len(entry_course_codes) + 8, column=2, columnspan=1, pady=10)
        text_result.grid(row=len(entry_course_codes) + 7, column=0, columnspan=5, pady=10)
        button_get_login_url.grid(row=len(entry_course_codes) + 8, column=1, columnspan=1, pady=10)
        button_get_logout_url.grid(row=len(entry_course_codes) + 8, column=3, columnspan=1, pady=5)




root = tk.Tk()
root.title("西财教务系统脚本")
style = Style(theme="flatly")
window = style.master
root.grid_columnconfigure(0, weight=1)
root.grid_rowconfigure(0, weight=1)
root.protocol("WM_DELETE_WINDOW", on_close)
main_frame = ttk.Frame(root)
main_frame.grid(row=0, column=0, padx=10, pady=10)
button_load_codes = ttk.Button(main_frame, text="填充选课代码", command=load_selection_codes,bootstyle="outline-info")
button_load_codes.grid(row=4, column=0, pady=5)
label_server_script_url = ttk.Label(main_frame, text="本专业选课界面URL：")
entry_server_script_url = ttk.Entry(main_frame, width=40)
label_code = ttk.Label(main_frame, text="选课代码(参考全校课表)：")
label_textbook_ordered = tk.Label(main_frame, text="是否订购教材：")
radio_var = tk.StringVar(value="0")
radio_button_yes = ttk.Radiobutton(main_frame, text="是", variable=radio_var, value="1")
radio_button_no = ttk.Radiobutton(main_frame, text="否", variable=radio_var, value="0")

entry_course_codes = [ttk.Entry(main_frame, width=40)]
entry_course_codes[0].grid(row=3, column=1, padx=10, pady=5, sticky="w")

entry_course_VIEWSTATE = [ttk.Entry(main_frame, width=30)]
entry_course_VIEWSTATE[0].grid(row=3, column=3, padx=10, pady=5, sticky="w")
label_VIEWSTATE = ttk.Label(main_frame, text="VIEWSTATE(留空则使用全局)")
label_VIEWSTATE.grid(row=2, column=3, padx=10, pady=5, sticky="w")

button_start_stop_monitor = ttk.Button(main_frame, text="启动监听", command=start_stop_monitor,bootstyle="outline-warning")
button_start_stop_monitor.grid(row=len(entry_course_codes) + 9, column=0, columnspan=1, pady=10)
button_add_entry = ttk.Button(main_frame, text="添加选课代码", command=add_course_entry,bootstyle="outline-info")
button_add_entry.grid(row=3, column=2, padx=10, pady=5)
button_remove_entry = ttk.Button(main_frame, text="删除选课代码", command=remove_course_entry,bootstyle="outline-info")
button_remove_entry.grid(row=4, column=2, padx=10, pady=5)
button_submit = ttk.Button(main_frame, text="提交", command=submit_form,bootstyle="outline-info")
button_submit.grid(row=len(entry_course_codes) + 5, column=1, columnspan=2, pady=10)
button_get_login_url = ttk.Button(main_frame, text="登录", command=get_login_url,bootstyle="outline-success")
button_get_login_url.grid(row=len(entry_course_codes) + 9, column=1, columnspan=1, pady=10)

button_get_logout_url = ttk.Button(main_frame, text="退出登录", command=clearlogin,bootstyle="outline-danger")
button_get_logout_url.grid(row=len(entry_course_codes) + 9, column=3, columnspan=1, pady=5)

button_config_window = ttk.Button(main_frame, text="修改配置文件", command=open_config_window,bootstyle="outline-dark")
button_config_window.grid(row=5, column=2, pady=5)
button_get_selection = ttk.Button(main_frame, text="获取当前选课情况", command=get_current_selection,bootstyle="outline-info")
button_get_selection.grid(row=len(entry_course_codes) + 9, column=2, columnspan=1, pady=10)
label_custom_interval = ttk.Label(main_frame, text="抢课间隔（秒）：")
custom_interval = ttk.Entry(main_frame, width=10)
custom_interval.insert(tk.END, "1")
button_edit_selection = ttk.Button(main_frame, text="编辑选课代码", command=open_selection_editor,bootstyle="outline-dark")
button_edit_selection.grid(row= 6, column=2, pady=5)
button_auto_submit = ttk.Button(main_frame, text="自动抢课", command=toggle_auto_submitting,bootstyle="outline-prime")
button_auto_fill = ttk.Button(main_frame, text="自动填充", command=auto_fill,bootstyle="outline-info")

text_result = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD, height=15, width=110, state=tk.DISABLED)
text_result.grid(row=7, column=0, columnspan=5, pady=10)

label_server_script_url.grid(row=0, column=0, padx=10, pady=5, sticky="w")
entry_server_script_url.grid(row=0, column=1, padx=10, pady=5, sticky="w")
label_code.grid(row=3, column=0, padx=5, pady=1, sticky='w')
label_textbook_ordered.grid(row=1, column=0, padx=10, pady=5, sticky="w")
radio_button_yes.grid(row=1, column=1, padx=10, pady=5, sticky="w")
radio_button_no.grid(row=1, column=2, padx=10, pady=5, sticky="w")
button_auto_fill.grid(row=0, column=2, padx=10, pady=5)

label_custom_interval.grid(row=5, column=0, padx=10, pady=5, sticky="w")
custom_interval.grid(row=5, column=1, padx=10, pady=5, sticky="w")
button_auto_submit.grid(row=6, column=0, columnspan=2, pady=5)
open_config_window_if_empty()
root.mainloop()
